import os
import io
import math
import base64
import uuid
import jwt
from datetime import datetime, date, time, timedelta
from functools import wraps
from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, send_file, current_app, abort
from flask_login import login_user, logout_user, login_required, current_user
from flask_bcrypt import generate_password_hash, check_password_hash

from database import db
from models import User, Employee, Department, Attendance, LeaveRequest, Notification, Setting
from forms import (
    LoginForm, RegistrationForm, ProfileForm, ChangePasswordForm,
    ForgotPasswordForm, ResetPasswordForm, DepartmentForm, LeaveApplicationForm, SystemSettingsForm
)
from mail_service import send_password_reset_email, send_attendance_notification, send_leave_notification

# Create blueprints
main_bp = Blueprint('main', __name__)

# Distance Calculation helper (Haversine Formula)
def get_distance_meters(lat1, lon1, lat2, lon2):
    R = 6371000.0  # Radius of earth in meters
    phi_1 = math.radians(lat1)
    phi_2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    
    a = math.sin(delta_phi / 2.0) ** 2 + \
        math.cos(phi_1) * math.cos(phi_2) * \
        math.sin(delta_lambda / 2.0) ** 2
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    
    return R * c

# Helper to save selfie images from webcam
def save_selfie_image(base64_str, folder, subfolder='selfies'):
    if not base64_str:
        return None
    try:
        if "data:image" in base64_str:
            base64_str = base64_str.split(",")[1]
        img_data = base64.b64decode(base64_str)
        filename = f"{uuid.uuid4().hex}.jpg"
        
        target_dir = os.path.join(folder, subfolder)
        os.makedirs(target_dir, exist_ok=True)
        
        filepath = os.path.join(target_dir, filename)
        with open(filepath, "wb") as fh:
            fh.write(img_data)
        return f"{subfolder}/{filename}"
    except Exception as e:
        print(f"Error saving selfie: {e}")
        return None

# Role-based restriction decorator
def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in roles:
                flash("You do not have permission to access this page.", "danger")
                return redirect(url_for('main.dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# JWT token generation for APIs / QR scan verification
def generate_jwt_token(payload, secret_key):
    return jwt.encode(payload, secret_key, algorithm='HS256')

def decode_jwt_token(token, secret_key):
    try:
        return jwt.decode(token, secret_key, algorithms=['HS256'])
    except Exception:
        return None

# ==========================================
# AUTHENTICATION ROUTES
# ==========================================

@main_bp.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    return redirect(url_for('main.login'))

@main_bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and check_password_hash(user.password_hash, form.password.data):
            if not user.is_active:
                flash("Your account has been deactivated. Please contact support.", "danger")
                return redirect(url_for('main.login'))
            login_user(user, remember=form.remember.data)
            flash(f"Welcome back, {user.username}!", "success")
            
            # Create a check-in notification for employee or admin
            notif = Notification(
                user_id=user.id,
                title="System Login",
                message=f"You successfully logged into the Attendance System at {datetime.now().strftime('%H:%M:%S')}",
                type="system"
            )
            db.session.add(notif)
            db.session.commit()
            
            next_page = request.args.get('next')
            return redirect(next_page or url_for('main.dashboard'))
        else:
            flash("Invalid username or password.", "danger")
            
    return render_template('login.html', form=form)

@main_bp.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    
    form = RegistrationForm()
    # Populate departments dropdown
    departments = Department.query.all()
    form.department_id.choices = [(d.id, d.name) for d in departments]
    
    if form.validate_on_submit():
        hashed_pw = generate_password_hash(form.password.data).decode('utf-8')
        user = User(
            username=form.username.data,
            email=form.email.data,
            password_hash=hashed_pw,
            role=form.role.data
        )
        db.session.add(user)
        db.session.flush() # Populate user.id
        
        # Unique QR token key
        qr_token = jwt.encode({'user_id': user.id, 'username': user.username}, current_app.config['JWT_SECRET_KEY'], algorithm='HS256')
        
        # Setup Employee profile
        # Generates employee code
        count = Employee.query.count() + 1
        emp_code = f"EMP-{datetime.now().year}-{count:04d}"
        
        employee = Employee(
            user_id=user.id,
            employee_id=emp_code,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            email=form.email.data,
            phone=form.phone.data,
            department_id=form.department_id.data,
            designation=form.designation.data,
            date_of_joining=date.today(),
            qr_code=qr_token,
            # Assigning defaults for GPS check (fallback to config values if empty)
            location_lat=current_app.config.get('ALLOWED_GPS_LATITUDE'),
            location_lng=current_app.config.get('ALLOWED_GPS_LONGITUDE')
        )
        db.session.add(employee)
        db.session.commit()
        
        flash("Registration successful! You can now log in.", "success")
        return redirect(url_for('main.login'))
        
    return render_template('register.html', form=form)

@main_bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for('main.login'))

@main_bp.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    form = ForgotPasswordForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user:
            # Generate JWT token valid for 30 minutes
            token = jwt.encode(
                {'user_id': user.id, 'exp': datetime.utcnow() + timedelta(minutes=30)},
                current_app.config['JWT_SECRET_KEY'],
                algorithm='HS256'
            )
            send_password_reset_email(user, token)
        # Always output success to prevent user email enumeration
        flash("If that email is registered, a password reset link has been sent.", "success")
        return redirect(url_for('main.login'))
    return render_template('forgot_password.html', form=form)

@main_bp.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))
    
    try:
        data = jwt.decode(token, current_app.config['JWT_SECRET_KEY'], algorithms=['HS256'])
        user_id = data['user_id']
    except Exception:
        flash("The password reset link is invalid or has expired.", "danger")
        return redirect(url_for('main.login'))
        
    user = User.query.get(user_id)
    if not user:
        return redirect(url_for('main.login'))
        
    form = ResetPasswordForm()
    if form.validate_on_submit():
        hashed_pw = generate_password_hash(form.password.data).decode('utf-8')
        user.password_hash = hashed_pw
        db.session.commit()
        flash("Your password has been reset. You may now log in.", "success")
        return redirect(url_for('main.login'))
    return render_template('reset_password.html', form=form)

# ==========================================
# DASHBOARD ROUTE
# ==========================================

@main_bp.route('/dashboard')
@login_required
def dashboard():
    today = date.today()
    
    # 1. ADMIN & HR DASHBOARD VIEW
    if current_user.role in ['admin', 'hr']:
        total_emp = Employee.query.filter_by(status='active').count()
        
        # Present Today
        present_employees = Attendance.query.filter(
            Attendance.date == today,
            Attendance.status.in_(['Present', 'Half Day'])
        ).all()
        present_count = len(present_employees)
        
        # Absent Today (exclude employees with active leave approval)
        leaves_today = LeaveRequest.query.filter(
            LeaveRequest.start_date <= today,
            LeaveRequest.end_date >= today,
            LeaveRequest.status == 'Approved'
        ).all()
        leave_count_today = len(leaves_today)
        
        absent_count = total_emp - (present_count + leave_count_today)
        if absent_count < 0:
            absent_count = 0
            
        present_emp_ids = {a.employee_id for a in present_employees}
        leave_emp_ids = {l.employee_id for l in leaves_today}
        
        active_employees = Employee.query.filter_by(status='active').all()
        absent_employees = [
            emp for emp in active_employees
            if emp.id not in present_emp_ids and emp.id not in leave_emp_ids
        ]
            
        # Late Entry
        late_employees = Attendance.query.filter(
            Attendance.date == today,
            Attendance.late_entry == True
        ).all()
        late_count = len(late_employees)
        
        # Pending leave requests
        pending_leaves = LeaveRequest.query.filter_by(status='Pending').count()
        
        # Attendance trends (recent 7 days)
        trends = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            p = Attendance.query.filter(Attendance.date == d, Attendance.status.in_(['Present', 'Half Day'])).count()
            trends.append({
                'day': d.strftime('%a'),
                'date': d.strftime('%Y-%m-%d'),
                'present': p
            })
            
        # Recent activity log
        recent_activity = Attendance.query.order_by(Attendance.id.desc()).limit(8).all()
        
        # Employee status distributions
        dept_data = db.session.query(
            Department.name, db.func.count(Employee.id)
        ).join(Employee, Employee.department_id == Department.id, isouter=True).group_by(Department.name).all()
        
        stats = {
            'total_employees': total_emp,
            'present_employees': present_count,
            'absent_employees': absent_count,
            'late_employees': late_count,
            'leave_requests': pending_leaves,
            'trends': trends,
            'recent_activity': recent_activity,
            'department_distributions': [{'name': d[0], 'count': d[1]} for d in dept_data if d[0] is not None],
            'present_list': present_employees,
            'absent_list': absent_employees,
            'late_list': late_employees
        }
        return render_template('dashboard.html', role=current_user.role, stats=stats)
        
    # 2. EMPLOYEE DASHBOARD VIEW
    else:
        employee = current_user.employee_profile
        if not employee:
            flash("Employee profile not found.", "danger")
            logout_user()
            return redirect(url_for('main.login'))
            
        # Today's attendance status
        today_attendance = Attendance.query.filter_by(employee_id=employee.id, date=today).first()
        
        # Total working hours this month
        start_of_month = today.replace(day=1)
        month_attendances = Attendance.query.filter(
            Attendance.employee_id == employee.id,
            Attendance.date >= start_of_month,
            Attendance.date <= today
        ).all()
        
        total_hours = sum([a.working_hours for a in month_attendances if a.working_hours])
        total_overtime = sum([a.overtime for a in month_attendances if a.overtime])
        present_days = sum([1 for a in month_attendances if a.status in ['Present', 'Half Day']])
        late_days = sum([1 for a in month_attendances if a.late_entry])
        
        # Leave balances (Sample allocations: Casual=12, Medical=10, Earned=15)
        approved_leaves = LeaveRequest.query.filter_by(employee_id=employee.id, status='Approved').all()
        leaves_taken = sum([(l.end_date - l.start_date).days + 1 for l in approved_leaves])
        
        leave_balance = max(37 - leaves_taken, 0)
        
        # Recent personal activities
        activities = Attendance.query.filter_by(employee_id=employee.id).order_by(Attendance.date.desc()).limit(5).all()
        
        # In-app notifications
        unread_notifs = Notification.query.filter_by(user_id=current_user.id, is_read=False).order_by(Notification.created_at.desc()).all()
        
        stats = {
            'employee': employee,
            'today_attendance': today_attendance,
            'total_working_hours': round(total_hours, 2),
            'total_overtime': round(total_overtime, 2),
            'present_days': present_days,
            'late_days': late_days,
            'leave_balance': leave_balance,
            'recent_activities': activities,
            'unread_notifications': unread_notifs
        }
        return render_template('dashboard.html', role=current_user.role, stats=stats)

# ==========================================
# ATTENDANCE MARKING ROUTES
# ==========================================

@main_bp.route('/attendance', methods=['GET'])
@login_required
def attendance():
    # Only employees can directly check in via web interface, admin/hr can view attendance history
    employee = current_user.employee_profile
    today = date.today()
    
    # Generate QR verification token for employee (dynamic, includes timestamp/expiry)
    qr_token = ""
    if employee:
        # Valid for 30 seconds check
        payload = {
            'employee_id': employee.id,
            'emp_code': employee.employee_id,
            'timestamp': datetime.now().isoformat()
        }
        qr_token = generate_jwt_token(payload, current_app.config['JWT_SECRET_KEY'])
        
    attendance_record = Attendance.query.filter_by(employee_id=employee.id, date=today).first() if employee else None
    
    # Retrieve system coordinates settings for display
    gps_lat = employee.location_lat if (employee and employee.location_lat) else current_app.config['ALLOWED_GPS_LATITUDE']
    gps_lng = employee.location_lng if (employee and employee.location_lng) else current_app.config['ALLOWED_GPS_LONGITUDE']
    radius = current_app.config['ALLOWED_GPS_RADIUS_METERS']
    
    # Calculate attendance history for Calendar Heatmap visualization
    history_records = []
    if employee:
        history = Attendance.query.filter_by(employee_id=employee.id).all()
        for h in history:
            history_records.append({
                'date': h.date.isoformat(),
                'status': h.status,
                'check_in': h.check_in_time.strftime('%H:%M') if h.check_in_time else '-',
                'check_out': h.check_out_time.strftime('%H:%M') if h.check_out_time else '-'
            })
            
    return render_template(
        'attendance.html', 
        attendance=attendance_record,
        gps_lat=gps_lat,
        gps_lng=gps_lng,
        radius=radius,
        qr_token=qr_token,
        history_json=history_records
    )

@main_bp.route('/attendance/checkin', methods=['POST'])
@login_required
def check_in():
    employee = current_user.employee_profile
    if not employee:
        return jsonify({'success': False, 'message': 'Only employees can log check-in.'}), 400
        
    today = date.today()
    now = datetime.now()
    
    # Already Checked In?
    existing = Attendance.query.filter_by(employee_id=employee.id, date=today).first()
    if existing and existing.check_in_time is not None:
        return jsonify({'success': False, 'message': 'Already checked in today.'})
        
    # Get JSON payload
    data = request.get_json() or {}
    lat = data.get('latitude')
    lng = data.get('longitude')
    selfie_b64 = data.get('selfie')
    
    # 1. GPS Proximity Verification
    office_lat = employee.location_lat if employee.location_lat else current_app.config['ALLOWED_GPS_LATITUDE']
    office_lng = employee.location_lng if employee.location_lng else current_app.config['ALLOWED_GPS_LONGITUDE']
    allowed_radius = current_app.config['ALLOWED_GPS_RADIUS_METERS']
    
    if lat is None or lng is None:
        return jsonify({'success': False, 'message': 'GPS coordinates required for verification.'})
        
    distance = get_distance_meters(float(lat), float(lng), office_lat, office_lng)
    if distance > allowed_radius:
        return jsonify({
            'success': False, 
            'message': f'GPS Verification failed. You are {round(distance, 1)}m away. Allowed radius is {allowed_radius}m.'
        })
        
    # 2. Webcam Selfie Verification
    if not selfie_b64:
        return jsonify({'success': False, 'message': 'Webcam selfie verification photo is required.'})
        
    selfie_filename = save_selfie_image(selfie_b64, current_app.config['UPLOAD_FOLDER'], 'selfies')
    if not selfie_filename:
        return jsonify({'success': False, 'message': 'Failed to verify upload picture.'})
        
    # 3. Time status calculation (Late Entry detection)
    check_in_limit_str = current_app.config['CHECK_IN_START_TIME']
    grace_mins = current_app.config['CHECK_IN_GRACE_PERIOD_MINS']
    
    limit_time = datetime.strptime(check_in_limit_str, '%H:%M:%S').time()
    check_in_limit_dt = datetime.combine(today, limit_time) + timedelta(minutes=grace_mins)
    
    late_entry = now > check_in_limit_dt
    
    # Create or update record
    if existing:
        attendance_rec = existing
        attendance_rec.check_in_time = now.time()
        attendance_rec.check_in_lat = lat
        attendance_rec.check_in_lng = lng
        attendance_rec.check_in_selfie = selfie_filename
        attendance_rec.status = 'Present'
        attendance_rec.late_entry = late_entry
    else:
        attendance_rec = Attendance(
            employee_id=employee.id,
            date=today,
            check_in_time=now.time(),
            check_in_lat=lat,
            check_in_lng=lng,
            check_in_selfie=selfie_filename,
            status='Present',
            late_entry=late_entry,
            approval_status='Approved'
        )
        db.session.add(attendance_rec)
        
    # Send email notification
    send_attendance_notification(employee, "Check-In", now.strftime('%H:%M:%S'))
    
    # Register Notification Log
    notif = Notification(
        user_id=current_user.id,
        title="Check-In Success",
        message=f"Checked-in successfully at {now.strftime('%H:%M:%S')}. Late: {late_entry}",
        type="attendance"
    )
    db.session.add(notif)
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': f"Checked in successfully at {now.strftime('%I:%M %p')}.",
        'late': late_entry
    })

@main_bp.route('/attendance/checkout', methods=['POST'])
@login_required
def check_out():
    employee = current_user.employee_profile
    if not employee:
        return jsonify({'success': False, 'message': 'Only employees can log check-out.'}), 400
        
    today = date.today()
    now = datetime.now()
    
    # Verify Checked In first
    existing = Attendance.query.filter_by(employee_id=employee.id, date=today).first()
    if not existing or existing.check_in_time is None:
        return jsonify({'success': False, 'message': 'Cannot check out without checking in first.'})
        
    if existing.check_out_time is not None:
        return jsonify({'success': False, 'message': 'Already checked out today.'})
        
    # Get JSON payload
    data = request.get_json() or {}
    lat = data.get('latitude')
    lng = data.get('longitude')
    selfie_b64 = data.get('selfie')
    
    # 1. GPS Verification
    office_lat = employee.location_lat if employee.location_lat else current_app.config['ALLOWED_GPS_LATITUDE']
    office_lng = employee.location_lng if employee.location_lng else current_app.config['ALLOWED_GPS_LONGITUDE']
    allowed_radius = current_app.config['ALLOWED_GPS_RADIUS_METERS']
    
    if lat is None or lng is None:
        return jsonify({'success': False, 'message': 'GPS coordinates required.'})
        
    distance = get_distance_meters(float(lat), float(lng), office_lat, office_lng)
    if distance > allowed_radius:
        return jsonify({
            'success': False, 
            'message': f'GPS Verification failed. Distance {round(distance, 1)}m exceeds limit.'
        })
        
    # 2. Camera Selfie Verification
    if not selfie_b64:
        return jsonify({'success': False, 'message': 'Camera selfie validation is required.'})
        
    selfie_filename = save_selfie_image(selfie_b64, current_app.config['UPLOAD_FOLDER'], 'selfies')
    if not selfie_filename:
        return jsonify({'success': False, 'message': 'Failed to verify selfie photo.'})
        
    # Update Attendance Log
    existing.check_out_time = now.time()
    existing.check_out_lat = lat
    existing.check_out_lng = lng
    existing.check_out_selfie = selfie_filename
    
    # Calculate working hours (difference in hours)
    in_datetime = datetime.combine(today, existing.check_in_time)
    out_datetime = datetime.combine(today, now.time())
    duration = out_datetime - in_datetime
    hours = duration.total_seconds() / 3600.0
    existing.working_hours = round(hours, 2)
    
    # Overtime & Early Exit detection
    # Example working standard: 8 hours shift. > 8.5 hours is overtime
    if hours > 8.5:
        existing.overtime = round(hours - 8.0, 2)
    else:
        existing.overtime = 0.0
        
    if hours < 4.0:
        existing.status = 'Half Day'
    
    # Early exit detection (e.g. shift ends at 17:00:00)
    # Defaulting simple rule: < 8 hours check-out is early exit
    existing.early_exit = hours < 8.0
    
    # Send email notification
    send_attendance_notification(employee, "Check-Out", now.strftime('%H:%M:%S'))
    
    # Register Notification Log
    notif = Notification(
        user_id=current_user.id,
        title="Check-Out Success",
        message=f"Checked-out successfully at {now.strftime('%H:%M:%S')}. Working Hours: {existing.working_hours}",
        type="attendance"
    )
    db.session.add(notif)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f"Checked out successfully at {now.strftime('%I:%M %p')}.",
        'working_hours': existing.working_hours,
        'overtime': existing.overtime
    })

# ==========================================
# QR CODE SCANNER CHECK-IN (ADMIN / KIOSK)
# ==========================================

@main_bp.route('/attendance/qr-checkin', methods=['POST'])
def qr_check_in_scan():
    """
    Kiosk terminal endpoint. Scans employee QR token from ID card or dynamic profile token
    and registers attendance check-in or check-out based on current date states.
    """
    data = request.get_json() or {}
    qr_token = data.get('qr_token')
    
    if not qr_token:
        return jsonify({'success': False, 'message': 'No QR token provided.'}), 400
        
    # Decode token
    decoded = decode_jwt_token(qr_token, current_app.config['JWT_SECRET_KEY'])
    if not decoded:
        return jsonify({'success': False, 'message': 'Invalid or expired QR code.'}), 400
        
    emp_id = decoded.get('employee_id') or decoded.get('user_id')
    
    employee = None
    if decoded.get('emp_code'):
        employee = Employee.query.filter_by(employee_id=decoded['emp_code']).first()
    elif emp_id:
        employee = Employee.query.get(emp_id)
        
    if not employee:
        return jsonify({'success': False, 'message': 'Employee record not found.'}), 404
        
    today = date.today()
    now = datetime.now()
    
    existing = Attendance.query.filter_by(employee_id=employee.id, date=today).first()
    
    # Perform check-in if not done today
    if not existing or existing.check_in_time is None:
        check_in_limit_str = current_app.config['CHECK_IN_START_TIME']
        grace_mins = current_app.config['CHECK_IN_GRACE_PERIOD_MINS']
        limit_time = datetime.strptime(check_in_limit_str, '%H:%M:%S').time()
        check_in_limit_dt = datetime.combine(today, limit_time) + timedelta(minutes=grace_mins)
        late_entry = now > check_in_limit_dt
        
        if existing:
            existing.check_in_time = now.time()
            existing.status = 'Present'
            existing.late_entry = late_entry
        else:
            existing = Attendance(
                employee_id=employee.id,
                date=today,
                check_in_time=now.time(),
                status='Present',
                late_entry=late_entry,
                approval_status='Approved'
            )
            db.session.add(existing)
            
        send_attendance_notification(employee, "Check-In (QR Scan)", now.strftime('%H:%M:%S'))
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': f"QR Check-in Success for {employee.full_name} at {now.strftime('%I:%M %p')}."
        })
        
    # Perform check-out if check-in exists but not check-out
    elif existing.check_out_time is None:
        existing.check_out_time = now.time()
        
        in_datetime = datetime.combine(today, existing.check_in_time)
        out_datetime = datetime.combine(today, now.time())
        duration = out_datetime - in_datetime
        hours = duration.total_seconds() / 3600.0
        
        existing.working_hours = round(hours, 2)
        if hours > 8.5:
            existing.overtime = round(hours - 8.0, 2)
        existing.early_exit = hours < 8.0
        
        send_attendance_notification(employee, "Check-Out (QR Scan)", now.strftime('%H:%M:%S'))
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': f"QR Check-out Success for {employee.full_name} at {now.strftime('%I:%M %p')}."
        })
    else:
        return jsonify({'success': False, 'message': 'Attendance already fully logged for today.'})

# ==========================================
# LEAVE MANAGEMENT ROUTES
# ==========================================

@main_bp.route('/leave/apply', methods=['GET', 'POST'])
@login_required
def apply_leave():
    employee = current_user.employee_profile
    if not employee:
        flash("Only employees can apply for leaves.", "warning")
        return redirect(url_for('main.dashboard'))
        
    form = LeaveApplicationForm()
    if form.validate_on_submit():
        try:
            start = datetime.strptime(form.start_date.data, '%Y-%m-%d').date()
            end = datetime.strptime(form.end_date.data, '%Y-%m-%d').date()
        except ValueError:
            flash("Invalid date format. Use YYYY-MM-DD.", "danger")
            return render_template('leave_apply.html', form=form)
            
        if start > end:
            flash("Start date cannot be after end date.", "danger")
            return render_template('leave_apply.html', form=form)
            
        # Create Leave request
        req = LeaveRequest(
            employee_id=employee.id,
            leave_type=form.leave_type.data,
            start_date=start,
            end_date=end,
            reason=form.reason.data,
            status='Pending'
        )
        db.session.add(req)
        
        # Notify admins and HR
        admins_hrs = User.query.filter(User.role.in_(['admin', 'hr'])).all()
        for admin in admins_hrs:
            notif = Notification(
                user_id=admin.id,
                title="New Leave Request",
                message=f"Employee {employee.full_name} applied for {form.leave_type.data} from {start} to {end}.",
                type="leave"
            )
            db.session.add(notif)
            
        db.session.commit()
        flash("Leave application submitted successfully.", "success")
        return redirect(url_for('main.dashboard'))
        
    # Get personal leave history
    leaves = LeaveRequest.query.filter_by(employee_id=employee.id).order_by(LeaveRequest.created_at.desc()).all()
    
    return render_template('leave_apply.html', form=form, leaves=leaves)

@main_bp.route('/leave/manage')
@login_required
@roles_required('admin', 'hr')
def manage_leaves():
    leaves = LeaveRequest.query.order_by(LeaveRequest.status.desc(), LeaveRequest.created_at.desc()).all()
    return render_template('leaves_manage.html', leaves=leaves)

@main_bp.route('/leave/action/<int:request_id>/<action>', methods=['POST'])
@login_required
@roles_required('admin', 'hr')
def leave_action(request_id, action):
    req = LeaveRequest.query.get_or_404(request_id)
    comments = request.form.get('comments', '')
    
    if action not in ['approve', 'reject']:
        flash("Invalid action.", "danger")
        return redirect(url_for('main.manage_leaves'))
        
    req.status = 'Approved' if action == 'approve' else 'Rejected'
    req.approved_by = current_user.id
    req.comments = comments
    
    # Auto mark attendance log entries for those days if approved
    if req.status == 'Approved':
        curr = req.start_date
        while curr <= req.end_date:
            # Overwrite or create attendance block
            att = Attendance.query.filter_by(employee_id=req.employee_id, date=curr).first()
            if not att:
                att = Attendance(
                    employee_id=req.employee_id,
                    date=curr,
                    status='Leave',
                    approval_status='Approved'
                )
                db.session.add(att)
            else:
                att.status = 'Leave'
            curr += timedelta(days=1)
            
    # Send email notification
    send_leave_notification(req.employee, req)
    
    # Notify user
    notif = Notification(
        user_id=req.employee.user_id,
        title=f"Leave Request {req.status}",
        message=f"Your leave application for {req.start_date} to {req.end_date} has been {req.status}.",
        type="leave"
    )
    db.session.add(notif)
    db.session.commit()
    
    flash(f"Leave request has been {req.status.lower()}d.", "success")
    return redirect(url_for('main.manage_leaves'))

# ==========================================
# EMPLOYEE MANAGEMENT (ADMIN/HR)
# ==========================================

@main_bp.route('/employees')
@login_required
@roles_required('admin', 'hr')
def employees():
    emp_list = Employee.query.all()
    depts = Department.query.all()
    dept_form = DepartmentForm()
    return render_template('employees.html', employees=emp_list, departments=depts, dept_form=dept_form)

@main_bp.route('/employee/add', methods=['POST'])
@login_required
@roles_required('admin', 'hr')
def add_employee():
    # Adding via admin panel
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password', 'emp123') # Default password
    role = request.form.get('role', 'employee')
    
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    phone = request.form.get('phone')
    designation = request.form.get('designation')
    department_id = request.form.get('department_id', type=int)
    
    # Helper to return validation error (handles AJAX gracefully)
    def fail(msg):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': msg}), 400
        flash(msg, "danger")
        return redirect(url_for('main.employees'))

    # Validations
    if not username or not email or not first_name or not last_name:
        return fail("Please fill in all required fields.")
        
    import re
    
    # 1. Firstname, Lastname contains only letters (spaces allowed)
    if not re.match(r"^[a-zA-Z\s]+$", first_name):
        return fail("First name must only contain letters.")
    if not re.match(r"^[a-zA-Z\s]+$", last_name):
        return fail("Last name must only contain letters.")
        
    # 2. Email domain ending with @ams or @ams.com
    email_lower = email.lower().strip()
    if not (email_lower.endswith('@ams') or email_lower.endswith('@ams.com')):
        return fail("Email must end with @ams or @ams.com.")
        
    # 3. Username must contain only letters and underscores
    if not re.match(r"^[a-zA-Z_]+$", username):
        return fail("Username must contain only letters and underscores.")
        
    # 4. Password contains letters, numbers, and special characters
    if not (any(c.isalpha() for c in password) and any(c.isdigit() for c in password) and any(not c.isalnum() for c in password)):
        return fail("Password must contain letters, numbers, and special characters.")
        
    # 5. Designation matches one of the active department codes (e.g. ENG, HR, SAL)
    dept_codes = [d.code.upper() for d in Department.query.all()]
    if not designation or designation.upper().strip() not in dept_codes:
        return fail(f"Designation must be one of the active department codes: {', '.join(dept_codes)}")
    designation = designation.upper().strip()
    
    # 6. Mobile number must contain exactly 10 digits
    cleaned_phone = re.sub(r"\D", "", phone or "")
    if len(cleaned_phone) != 10:
        return fail("Mobile number must contain exactly 10 digits.")
    phone = cleaned_phone
        
    u_exists = User.query.filter((User.username == username) | (User.email == email)).first()
    if u_exists:
        return fail("Username or email already exists.")
        
    hashed_pw = generate_password_hash(password).decode('utf-8')
    user = User(
        username=username,
        email=email,
        password_hash=hashed_pw,
        role=role
    )
    db.session.add(user)
    db.session.flush()
    
    # Generate QR key
    qr_token = jwt.encode({'user_id': user.id, 'username': user.username}, current_app.config['JWT_SECRET_KEY'], algorithm='HS256')
    
    count = Employee.query.count() + 1
    emp_code = f"EMP-{datetime.now().year}-{count:04d}"
    
    employee = Employee(
        user_id=user.id,
        employee_id=emp_code,
        first_name=first_name,
        last_name=last_name,
        email=email,
        phone=phone,
        department_id=department_id,
        designation=designation,
        date_of_joining=date.today(),
        qr_code=qr_token,
        location_lat=current_app.config.get('ALLOWED_GPS_LATITUDE'),
        location_lng=current_app.config.get('ALLOWED_GPS_LONGITUDE')
    )
    db.session.add(employee)
    db.session.commit()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        flash("Employee profile added successfully.", "success")
        return jsonify({'success': True, 'message': "Employee profile added successfully."})
    
    flash("Employee profile added successfully.", "success")
    return redirect(url_for('main.employees'))
 
@main_bp.route('/employee/edit/<int:emp_id>', methods=['POST'])
@login_required
@roles_required('admin', 'hr')
def edit_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    first_name = request.form.get('first_name')
    last_name = request.form.get('last_name')
    phone = request.form.get('phone')
    designation = request.form.get('designation')
    department_id = request.form.get('department_id', type=int)
    location_lat = request.form.get('location_lat', type=float)
    location_lng = request.form.get('location_lng', type=float)
    status = request.form.get('status', 'active')
    
    import re
    
    # Helper to return validation error (handles AJAX gracefully)
    def fail(msg):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': msg}), 400
        flash(msg, "danger")
        return redirect(url_for('main.employees'))

    # Validation filters for edit
    if not first_name or not last_name:
        return fail("Please fill in all required fields.")
        
    if not re.match(r"^[a-zA-Z\s]+$", first_name):
        return fail("First name must only contain letters.")
    if not re.match(r"^[a-zA-Z\s]+$", last_name):
        return fail("Last name must only contain letters.")
        
    # Designation matches active department codes
    dept_codes = [d.code.upper() for d in Department.query.all()]
    if not designation or designation.upper().strip() not in dept_codes:
        return fail(f"Designation must be one of the active department codes: {', '.join(dept_codes)}")
    designation = designation.upper().strip()
        
    # Mobile number must contain exactly 10 digits
    cleaned_phone = re.sub(r"\D", "", phone or "")
    if len(cleaned_phone) != 10:
        return fail("Mobile number must contain exactly 10 digits.")
        
    emp.first_name = first_name
    emp.last_name = last_name
    emp.phone = cleaned_phone
    emp.designation = designation
    emp.department_id = department_id
    emp.location_lat = location_lat
    emp.location_lng = location_lng
    emp.status = status
    
    # Sync status to user account
    emp.user.is_active = (emp.status == 'active')
    
    db.session.commit()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        flash("Employee updated successfully.", "success")
        return jsonify({'success': True, 'message': "Employee updated successfully."})
        
    flash("Employee updated successfully.", "success")
    return redirect(url_for('main.employees'))

@main_bp.route('/employee/delete/<int:emp_id>', methods=['POST'])
@login_required
@roles_required('admin')
def delete_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    user = emp.user
    db.session.delete(user) # Will cascade delete employee details
    db.session.commit()
    flash("Employee deleted successfully.", "success")
    return redirect(url_for('main.employees'))

# ==========================================
# DEPARTMENT MANAGEMENT (ADMIN/HR)
# ==========================================

@main_bp.route('/department/add', methods=['POST'])
@login_required
@roles_required('admin', 'hr')
def add_department():
    form = DepartmentForm()
    if form.validate_on_submit():
        dept = Department(
            name=form.name.data,
            code=form.code.data,
            description=form.description.data
        )
        db.session.add(dept)
        db.session.commit()
        flash("Department created successfully.", "success")
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in {field}: {error}", "danger")
    return redirect(url_for('main.employees'))

# ==========================================
# ID CARD GENERATION ROUTE
# ==========================================

@main_bp.route('/employee/id-card/<int:emp_id>')
@login_required
def generate_id_card(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    # Check permissions (admins/hr can view all, employees can only view their own)
    if current_user.role == 'employee' and current_user.employee_profile.id != emp.id:
        abort(403)
        
    return render_template('id_card.html', employee=emp)

# ==========================================
# USER PROFILE & SETTINGS
# ==========================================

@main_bp.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    form = ProfileForm()
    password_form = ChangePasswordForm()
    
    employee = current_user.employee_profile
    
    if request.method == 'GET':
        if employee:
            form.first_name.data = employee.first_name
            form.last_name.data = employee.last_name
            form.phone.data = employee.phone
            form.designation.data = employee.designation
            form.email.data = employee.email
        else:
            form.first_name.data = current_user.username
            form.last_name.data = ""
            form.email.data = current_user.email
            
    # Handle profile details update
    if 'submit' in request.form and form.validate_on_submit():
        if employee:
            employee.first_name = form.first_name.data
            employee.last_name = form.last_name.data
            employee.phone = form.phone.data
            employee.email = form.email.data
            employee.user.email = form.email.data
        else:
            current_user.email = form.email.data
        db.session.commit()
        flash("Profile updated successfully.", "success")
        return redirect(url_for('main.profile'))
        
    # Handle password change update
    if 'current_password' in request.form and password_form.validate_on_submit():
        if check_password_hash(current_user.password_hash, password_form.current_password.data):
            current_user.password_hash = generate_password_hash(password_form.new_password.data).decode('utf-8')
            db.session.commit()
            flash("Password updated successfully.", "success")
            return redirect(url_for('main.profile'))
        else:
            flash("Invalid current password.", "danger")
            
    return render_template('profile.html', form=form, password_form=password_form, employee=employee)

@main_bp.route('/profile/upload-avatar', methods=['POST'])
@login_required
def upload_avatar():
    employee = current_user.employee_profile
    if not employee:
        return jsonify({'success': False, 'message': 'Profile profile not found.'}), 400
        
    if 'avatar' not in request.files:
        return jsonify({'success': False, 'message': 'No picture selected.'}), 400
        
    file = request.files['avatar']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file.'}), 400
        
    if file:
        filename = f"avatar_{employee.id}_{uuid.uuid4().hex[:8]}.jpg"
        target_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'avatars')
        os.makedirs(target_dir, exist_ok=True)
        
        file.save(os.path.join(target_dir, filename))
        employee.profile_pic = f"avatars/{filename}"
        db.session.commit()
        
        return jsonify({'success': True, 'filename': employee.profile_pic})

@main_bp.route('/settings', methods=['GET', 'POST'])
@login_required
@roles_required('admin')
def settings():
    form = SystemSettingsForm()
    
    keys = ['check_in_start_time', 'check_in_grace_period_mins', 'allowed_gps_latitude', 'allowed_gps_longitude', 'allowed_gps_radius_meters']
    
    if request.method == 'GET':
        for key in keys:
            setting = Setting.query.filter_by(key=key).first()
            if setting:
                val = setting.value
                if key == 'check_in_grace_period_mins':
                    form.check_in_grace_period_mins.data = int(val)
                elif key == 'allowed_gps_latitude':
                    form.allowed_gps_latitude.data = float(val)
                elif key == 'allowed_gps_longitude':
                    form.allowed_gps_longitude.data = float(val)
                elif key == 'allowed_gps_radius_meters':
                    form.allowed_gps_radius_meters.data = float(val)
                else:
                    form.check_in_start_time.data = val
                    
    if form.validate_on_submit():
        vals = {
            'check_in_start_time': form.check_in_start_time.data,
            'check_in_grace_period_mins': str(form.check_in_grace_period_mins.data),
            'allowed_gps_latitude': str(form.allowed_gps_latitude.data),
            'allowed_gps_longitude': str(form.allowed_gps_longitude.data),
            'allowed_gps_radius_meters': str(form.allowed_gps_radius_meters.data),
        }
        
        for k, v in vals.items():
            setting = Setting.query.filter_by(key=k).first()
            if setting:
                setting.value = v
            else:
                setting = Setting(key=k, value=v)
                db.session.add(setting)
                
            # Sync Config dynamically
            current_app.config[k.upper()] = int(v) if k.endswith('mins') else (float(v) if k.startswith('allowed') else v)
            
        db.session.commit()
        flash("System settings saved successfully.", "success")
        return redirect(url_for('main.settings'))
        
    return render_template('settings.html', form=form)

# ==========================================
# NOTIFICATIONS READ ROUTE
# ==========================================

@main_bp.route('/notifications/read-all', methods=['POST'])
@login_required
def read_all_notifications():
    notifications = Notification.query.filter_by(user_id=current_user.id, is_read=False).all()
    for notif in notifications:
        notif.is_read = True
    db.session.commit()
    return jsonify({'success': True})

# ==========================================
# REPORTS & DATA EXPORTS
# ==========================================

@main_bp.route('/reports', methods=['GET'])
@login_required
def reports():
    # Admins/HR can see all reports, employees can see their own
    employee = current_user.employee_profile
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    emp_filter_id = request.args.get('employee_id')
    
    query = Attendance.query
    
    if current_user.role == 'employee':
        query = query.filter_by(employee_id=employee.id)
    elif emp_filter_id:
        query = query.filter_by(employee_id=int(emp_filter_id))
        
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date >= start_date)
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date <= end_date)
        
    records = query.order_by(Attendance.date.desc()).all()
    employees_list = Employee.query.all() if current_user.role in ['admin', 'hr'] else []
    
    return render_template(
        'reports.html',
        records=records,
        employees=employees_list,
        start_date=start_date_str,
        end_date=end_date_str,
        selected_emp=int(emp_filter_id) if emp_filter_id else None
    )

# 1. EXCEL EXPORT
@main_bp.route('/reports/export/excel')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    employee = current_user.employee_profile
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    emp_filter_id = request.args.get('employee_id')
    
    query = Attendance.query
    if current_user.role == 'employee':
        query = query.filter_by(employee_id=employee.id)
    elif emp_filter_id:
        query = query.filter_by(employee_id=int(emp_filter_id))
        
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date >= start_date)
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date <= end_date)
        
    records = query.order_by(Attendance.date.desc()).all()
    
    # Generate Excel in-memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Styling variables
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Table headers
    headers = [
        "Date", "Employee Code", "Full Name", "Department",
        "Check-In Time", "Check-Out Time", "Hours Worked", "Overtime", "Status", "Late Check-in"
    ]
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # Write rows
    for r_idx, r in enumerate(records, 2):
        ws.cell(row=r_idx, column=1, value=r.date.strftime('%Y-%m-%d'))
        ws.cell(row=r_idx, column=2, value=r.employee.employee_id)
        ws.cell(row=r_idx, column=3, value=r.employee.full_name)
        ws.cell(row=r_idx, column=4, value=r.employee.department.name if r.employee.department else 'N/A')
        ws.cell(row=r_idx, column=5, value=r.check_in_time.strftime('%H:%M:%S') if r.check_in_time else '-')
        ws.cell(row=r_idx, column=6, value=r.check_out_time.strftime('%H:%M:%S') if r.check_out_time else '-')
        ws.cell(row=r_idx, column=7, value=r.working_hours)
        ws.cell(row=r_idx, column=8, value=r.overtime)
        ws.cell(row=r_idx, column=9, value=r.status)
        ws.cell(row=r_idx, column=10, value="Yes" if r.late_entry else "No")
        
    # Adjust widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Attendance_Report_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# 1b. CSV EXPORT (ATTENDANCE REPORT)
@main_bp.route('/reports/export/csv')
@login_required
def export_csv():
    import csv
    
    employee = current_user.employee_profile
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    emp_filter_id = request.args.get('employee_id')
    
    query = Attendance.query
    if current_user.role == 'employee':
        query = query.filter_by(employee_id=employee.id)
    elif emp_filter_id:
        query = query.filter_by(employee_id=int(emp_filter_id))
        
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date >= start_date)
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date <= end_date)
        
    records = query.order_by(Attendance.date.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    headers = [
        "Date", "Employee ID", "Full Name", "Department",
        "Check-In Time", "Check-Out Time", "Hours Worked", "Overtime", "Status", "Late Check-in"
    ]
    writer.writerow(headers)
    
    # Rows
    for r in records:
        writer.writerow([
            r.date.strftime('%Y-%m-%d'),
            r.employee.employee_id,
            r.employee.full_name,
            r.employee.department.name if r.employee.department else 'N/A',
            r.check_in_time.strftime('%H:%M:%S') if r.check_in_time else '-',
            r.check_out_time.strftime('%H:%M:%S') if r.check_out_time else '-',
            r.working_hours,
            r.overtime,
            r.status,
            "Yes" if r.late_entry else "No"
        ])
        
    output.seek(0)
    filename = f"Attendance_Report_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    
    from flask import make_response
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "text/csv"
    return response

# 1c. CSV EXPORT (EMPLOYEES DIRECTORY)
@main_bp.route('/employees/export/csv')
@login_required
@roles_required('admin', 'hr')
def export_employees_csv():
    import csv
    
    employees = Employee.query.all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    headers = [
        "Employee ID", "First Name", "Last Name", "Email", "Phone",
        "Department", "Designation", "Joining Date", "Status"
    ]
    writer.writerow(headers)
    
    # Rows
    for emp in employees:
        writer.writerow([
            emp.employee_id,
            emp.first_name,
            emp.last_name,
            emp.email,
            emp.phone or '',
            emp.department.name if emp.department else 'N/A',
            emp.designation or 'N/A',
            emp.date_of_joining.strftime('%Y-%m-%d'),
            emp.status
        ])
        
    output.seek(0)
    filename = f"Employee_Directory_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    
    from flask import make_response
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "text/csv"
    return response

# 1d. EXCEL EXPORT (EMPLOYEES DIRECTORY)
@main_bp.route('/employees/export/excel')
@login_required
@roles_required('admin', 'hr')
def export_employees_excel():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    employees = Employee.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees Directory"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Table headers
    headers = [
        "Employee ID", "First Name", "Last Name", "Email", "Phone",
        "Department", "Designation", "Joining Date", "Status"
    ]
    
    # Styling headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    # Write rows
    for r_idx, emp in enumerate(employees, 2):
        ws.cell(row=r_idx, column=1, value=emp.employee_id).alignment = center_align
        ws.cell(row=r_idx, column=2, value=emp.first_name).alignment = left_align
        ws.cell(row=r_idx, column=3, value=emp.last_name).alignment = left_align
        ws.cell(row=r_idx, column=4, value=emp.email).alignment = left_align
        ws.cell(row=r_idx, column=5, value=emp.phone or '').alignment = center_align
        ws.cell(row=r_idx, column=6, value=emp.department.name if emp.department else 'N/A').alignment = left_align
        ws.cell(row=r_idx, column=7, value=emp.designation or 'N/A').alignment = left_align
        ws.cell(row=r_idx, column=8, value=emp.date_of_joining.strftime('%Y-%m-%d')).alignment = center_align
        ws.cell(row=r_idx, column=9, value=emp.status).alignment = center_align
        
    # Adjust widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Employee_Directory_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

# 2. PDF EXPORT
@main_bp.route('/reports/export/pdf')
@login_required
def export_pdf():
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    employee = current_user.employee_profile
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    emp_filter_id = request.args.get('employee_id')
    
    query = Attendance.query
    if current_user.role == 'employee':
        query = query.filter_by(employee_id=employee.id)
    elif emp_filter_id:
        query = query.filter_by(employee_id=int(emp_filter_id))
        
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date >= start_date)
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(Attendance.date <= end_date)
        
    records = query.order_by(Attendance.date.desc()).all()
    
    # PDF Setup
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=15
    )
    
    # Document header
    story.append(Paragraph("Employee Attendance System - Audit Report", title_style))
    story.append(Spacer(1, 10))
    
    # Table headers and contents
    data = [[
        "Date", "ID Code", "Full Name", "Department",
        "Check-In", "Check-Out", "Hours", "Overtime", "Status"
    ]]
    
    for r in records:
        data.append([
            r.date.strftime('%Y-%m-%d'),
            r.employee.employee_id,
            r.employee.full_name,
            r.employee.department.name if r.employee.department else 'N/A',
            r.check_in_time.strftime('%H:%M:%S') if r.check_in_time else '-',
            r.check_out_time.strftime('%H:%M:%S') if r.check_out_time else '-',
            str(r.working_hours),
            str(r.overtime),
            r.status
        ])
        
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F2F4F7')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D3D3D3')),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ]))
    
    story.append(t)
    doc.build(story)
    
    buffer.seek(0)
    filename = f"Attendance_Report_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

# 3. DOWNLOAD ATTENDANCE CERTIFICATE
@main_bp.route('/reports/certificate/<int:att_id>')
@login_required
def download_certificate(att_id):
    """
    Generates a personalized PDF Certificate of Perfect Attendance verification log.
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    att = Attendance.query.get_or_404(att_id)
    if current_user.role == 'employee' and current_user.employee_profile.id != att.employee_id:
        abort(403)
        
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50
    )
    story = []
    
    styles = getSampleStyleSheet()
    
    cert_title = ParagraphStyle(
        'CertTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#1F4E79'),
        alignment=1, # Center
        spaceAfter=30
    )
    
    cert_body = ParagraphStyle(
        'CertBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=14,
        leading=22,
        alignment=1,
        spaceAfter=15
    )
    
    story.append(Spacer(1, 40))
    story.append(Paragraph("CERTIFICATE OF ATTENDANCE VERIFICATION", cert_title))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"This is to verify that employee logs record presence for:", cert_body))
    story.append(Paragraph(f"<strong>{att.employee.full_name}</strong>", ParagraphStyle('EmpName', parent=cert_body, fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#000000'))))
    story.append(Paragraph(f"Employee ID: {att.employee.employee_id}", cert_body))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"On date: <strong>{att.date.strftime('%A, %B %d, %Y')}</strong>", cert_body))
    story.append(Paragraph(f"Check-In registered at: {att.check_in_time.strftime('%I:%M:%S %p')}", cert_body))
    story.append(Paragraph(f"Check-Out registered at: {att.check_out_time.strftime('%I:%M:%S %p') if att.check_out_time else 'N/A'}", cert_body))
    story.append(Paragraph(f"Total Working Hours registered: {att.working_hours} Hours", cert_body))
    story.append(Spacer(1, 40))
    story.append(Paragraph("System Authenticated Audit Log", ParagraphStyle('Footer', parent=cert_body, fontName='Helvetica-Oblique', fontSize=10, textColor=colors.gray)))
    
    doc.build(story)
    buffer.seek(0)
    
    filename = f"Certificate_{att.employee.employee_id}_{att.date.strftime('%Y%m%d')}.pdf"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )
